import pandas as pd
import numpy as np
from datetime import datetime
import os
import re
import warnings

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def compactar_tramos(tramos_str):
    if not isinstance(tramos_str, str):
        if pd.isna(tramos_str):
            return ""
        tramos_str = str(tramos_str)
        
    numeros = []
    for t in tramos_str.split(','):
        t = t.strip()
        if not t: continue
        try:
            numeros.append(float(t))
        except ValueError:
            pass
            
    if not numeros: return ""
    
    numeros = sorted(list(set(numeros)))
    if len(numeros) == 1:
        return str(int(numeros[0]) if numeros[0].is_integer() else numeros[0])
        
    rangos = []
    inicio = numeros[0]
    anterior = numeros[0]
    
    for num in numeros[1:]:
        if num == anterior + 1:
            anterior = num
        else:
            if inicio == anterior:
                rangos.append(str(int(inicio) if inicio.is_integer() else inicio))
            else:
                rangos.append(f"{int(inicio) if inicio.is_integer() else inicio}-{int(anterior) if anterior.is_integer() else anterior}")
            inicio = num
            anterior = num
            
    if inicio == anterior:
        rangos.append(str(int(inicio) if inicio.is_integer() else inicio))
    else:
        rangos.append(f"{int(inicio) if inicio.is_integer() else inicio}-{int(anterior) if anterior.is_integer() else anterior}")
        
    return ", ".join(rangos)

def buscar_archivo_contenedores():
    """Busca el único archivo Contenedores_AAAAMMDD_HHMM.ods en el directorio del script.
    Retorna (ruta_completa, datetime_archivo) o lanza un error si no se encuentra."""
    import glob
    base_dir = os.path.dirname(os.path.abspath(__file__))
    patron = os.path.join(base_dir, "Contenedores_*.ods")
    archivos = glob.glob(patron)
    
    if len(archivos) == 0:
        raise FileNotFoundError(
            f"No se encontró ningún archivo 'Contenedores_*.ods' en: {base_dir}"
        )
    if len(archivos) > 1:
        raise FileNotFoundError(
            f"Se encontraron múltiples archivos Contenedores_*.ods en {base_dir}: {archivos}\n"
            "Debe haber exactamente uno."
        )
    
    ruta = archivos[0]
    nombre = os.path.basename(ruta)  # e.g. Contenedores_20260415_1414.ods
    
    # Extraer fecha y hora del nombre: Contenedores_AAAAMMDD_HHMM.ods
    m = re.match(r'Contenedores_(\d{8})_(\d{4})\.ods', nombre)
    if not m:
        raise ValueError(
            f"El archivo '{nombre}' no tiene el formato esperado 'Contenedores_AAAAMMDD_HHMM.ods'"
        )
    
    fecha_str = m.group(1)  # '20260415'
    hora_str = m.group(2)   # '1414'
    dt_archivo = datetime.strptime(fecha_str + hora_str, "%Y%m%d%H%M")
    
    return ruta, dt_archivo


def load_data():
    ruta_archivo, dt_archivo = buscar_archivo_contenedores()
    
    try:
        df = pd.read_excel(ruta_archivo, engine='odf')
    except Exception as e:
        print(f"Error cargando {ruta_archivo}: {e}")
        return pd.DataFrame(), None
        
    # Convertir fechas
    for col in ['Fecha No Levante', 'Fecha Último Levante']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], format='%d/%m/%Y %H:%M', errors='coerce')
            
    # Asignar Oficina
    df['Oficina'] = np.where(df['Circuito'].astype(str).str.match(r'^B_0?[1-7](\b|$)'), 'Fideicomiso', 'IM')
    
    # Filtrar solo municipal
    df = df[df['Oficina'] == 'IM'].copy()
    
    # Extraer Municipio
    df['Municipio'] = df['Circuito'].astype(str).str.split('_').str[0]
    
    now = datetime.now()
    # Acumulaciones
    df['Acumulacion_horas'] = (now - df['Fecha Último Levante']).dt.total_seconds() / 3600
    df['Acumulacion_horas'] = df['Acumulacion_horas'].round(1)
    
    df['Acumulacion_dias_porhora'] = (now - df['Fecha Último Levante']).dt.total_seconds() / 86400
    df['Acumulacion_dias_porhora'] = df['Acumulacion_dias_porhora'].round(0)
    
    df['Acumulacion_dias_calendario'] = (pd.to_datetime(now.date()) - df['Fecha Último Levante'].dt.floor('D')).dt.days
    
    return df, dt_archivo
    
def procesar_atrasos(df):
    MOTIVOS_EXCLUIR_ATRASO = [
        "No Está (20)","Roto (18)","Fuego (19)","Sobrepeso (11)",
        "Fuera de Alcance (21)","Tapa Bloqueda (14)",
        "Buzonera Girada (24)","Volcado (22)"
    ]
    
    # Atrasos mayores a 3 dias que no tengan Fecha No levante cargada
    df_atrasos = df[(df['Acumulacion_dias_calendario'] >= 3) & (df['Fecha No Levante'].isna())].copy()
    df_atrasos = df_atrasos[~df_atrasos['Motivo No Levante'].isin(MOTIVOS_EXCLUIR_ATRASO)].copy()
    
    def categorizar(dias):
        if pd.isna(dias): return "Otros"
        d = int(dias)
        if d >= 6: return ">=6"
        elif d == 5: return "5"
        elif d == 4: return "4"
        elif d == 3: return "3"
        return "Otros"
        
    df_atrasos['Categoria_Atraso'] = df_atrasos['Acumulacion_dias_calendario'].apply(categorizar)
    df_atrasos = df_atrasos[df_atrasos['Categoria_Atraso'] != "Otros"]
    
    # TABLA 1: Resumen de Atrasos
    resumen = df_atrasos.groupby(['Municipio', 'Categoria_Atraso']).size().reset_index(name='conteo')
    pivot = resumen.pivot(index='Municipio', columns='Categoria_Atraso', values='conteo').fillna(0)
    
    columnas_orden = [">=6", "5", "4", "3"]
    for col in columnas_orden:
        if col not in pivot.columns:
            pivot[col] = 0
            
    pivot = pivot[columnas_orden].astype(int)
    pivot['Total'] = pivot.sum(axis=1)
    
    pivot = pivot.reset_index()
    total_gral = pivot.sum(numeric_only=True).to_dict()
    total_gral['Municipio'] = 'TOTAL'
    tabla_final = pd.concat([pivot, pd.DataFrame([total_gral])], ignore_index=True)
    
    # TABLA 2: Detalle de Atrasos
    # Agrupamos los datos manualmente en vez de custom pandas agg para tramos
    def agg_tramos(x):
        return compactar_tramos(','.join(x.dropna().astype(str)))
        
    detalle = df_atrasos.groupby(['Municipio', 'Categoria_Atraso', 'Circuito']).agg(
        cantidad=('Circuito', 'count'),
        tramos_posicion=('Posición', agg_tramos),
        fecha_levante=('Fecha Último Levante', lambda x: x.min().strftime('%d/%m/%y') if not pd.isna(x.min()) else "")
    ).reset_index()
    
    # Ordenar por Municipio -> Atraso -> Cantidad Desc
    orden_cat = {">=6": 0, "5": 1, "4": 2, "3": 3}
    detalle['sort_cat'] = detalle['Categoria_Atraso'].map(orden_cat)
    detalle = detalle.sort_values(['Municipio', 'sort_cat', 'cantidad'], ascending=[True, True, False])
    detalle = detalle.drop('sort_cat', axis=1)
    
    # Fila total
    tot_det = {'Municipio': 'TOTAL', 'Categoria_Atraso':'', 'Circuito':'', 'cantidad': detalle['cantidad'].sum(), 'tramos_posicion':'', 'fecha_levante':''}
    detalle = pd.concat([detalle, pd.DataFrame([tot_det])], ignore_index=True)
    
    detalle = detalle.rename(columns={
        "Categoria_Atraso": "Atraso",
        "cantidad": "Cant.",
        "tramos_posicion": "Tramos Posición",
        "fecha_levante": "Fecha Levante"
    })
    
    return tabla_final, detalle

def procesar_grua(df):
    MOTIVOS_GRUA = ["Roto (18)", "Sobrepeso (11)", "Fuera de Alcance (21)", "Buzonera Girada (24)", "Cruzado(23)", "Calle Cerrada(13)"]
    df_grua = df[df['Motivo No Levante'].isin(MOTIVOS_GRUA)].copy()
    
    def agg_tramos(x):
        return compactar_tramos(','.join(x.dropna().astype(str)))
        
    resumen = df_grua.groupby('Circuito').agg(
        suma_una=('% de acumulación (UNA)', 'sum'),
        Cant_=('Circuito', 'count'),
        tramos_posicion=('Posición', agg_tramos),
        fecha_ult_levante_mas_vieja=('Fecha Último Levante', lambda x: x.min().strftime('%d/%m/%y') if not pd.isna(x.min()) else "")
    ).reset_index()
    
    resumen = resumen.sort_values('suma_una', ascending=False)
    
    tot_gral = {
        'Circuito': 'TOTAL',
        'suma_una': '',
        'Cant_': resumen['Cant_'].sum(),
        'tramos_posicion': '',
        'fecha_ult_levante_mas_vieja': ''
    }
    resumen = pd.concat([resumen, pd.DataFrame([tot_gral])], ignore_index=True)
    resumen.columns = ["Circuito", "Suma UNA", "Cant.", "Tramos Posición", "Fecha Ult Levante"]
    
    listado = df_grua[['Circuito', 'Posición', 'Fecha Último Levante']].copy()
    listado = listado.sort_values(by=['Fecha Último Levante', 'Circuito', 'Posición'])
    listado['Fecha último levante'] = listado['Fecha Último Levante'].dt.strftime('%d/%m/%y')
    listado = listado.drop('Fecha Último Levante', axis=1)
    
    return resumen, listado

def procesar_fuego(df):
    df_f = df[df['Motivo No Levante'] == "Fuego (19)"].copy()
    
    def agg_tramos(x):
        return compactar_tramos(','.join(x.dropna().astype(str)))
        
    resumen = df_f.groupby('Circuito').agg(
        Cantidad=('Circuito', 'count'),
        tramos_posicion=('Posición', agg_tramos)
    ).reset_index()
    resumen = resumen.sort_values('Cantidad', ascending=False)
    
    tot = {'Circuito': 'TOTAL', 'Cantidad': resumen['Cantidad'].sum(), 'tramos_posicion': ''}
    resumen = pd.concat([resumen, pd.DataFrame([tot])], ignore_index=True)
    resumen.columns = ["Circuito", "Cantidad", "Tramos Posición"]
    
    return resumen
    
def procesar_no_esta(df):
    df_ne = df[df['Motivo No Levante'] == "No Está (20)"].copy()
    
    def agg_tramos(x):
        return compactar_tramos(','.join(x.dropna().astype(str)))
        
    resumen = df_ne.groupby('Circuito').agg(
        Cantidad=('Circuito', 'count'),
        tramos_posicion=('Posición', agg_tramos)
    ).reset_index()
    resumen = resumen.sort_values('Cantidad', ascending=False)
    
    tot = {'Circuito': 'TOTAL', 'Cantidad': resumen['Cantidad'].sum(), 'tramos_posicion': ''}
    resumen = pd.concat([resumen, pd.DataFrame([tot])], ignore_index=True)
    resumen.columns = ["Circuito", "Cantidad", "Tramos Posición"]
    
    
    # Detalle con el historico 
    try:
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        ruta_rds = os.path.join(base_dir, "db", "GOL_reportes", "historico_llenadoGol.rds")
        
        import pyreadr
        res = pyreadr.read_r(ruta_rds)
        hist = res[None].copy() # dataframe crudo
        
        # Filtro 90 días hacia atrás a partir de hoy
        fecha_corte = pd.to_datetime(datetime.now().date() - pd.Timedelta(days=90))
        hist['Fecha_hora_pasaje'] = pd.to_datetime(hist['Fecha_hora_pasaje'], errors='coerce')
        hist = hist[hist['Fecha_hora_pasaje'] >= fecha_corte].copy()
        
        # Normalizar para buscar coincidencias (como en R)
        def normalizar_txt(s):
            if pd.isna(s): return ""
            s = str(s).upper()
            import unicodedata
            s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
            s = re.sub(r'\s+', ' ', s).strip()
            return s
            
        hist['motivo_norm'] = hist['Incidencia'].apply(normalizar_txt)
        hist['gid'] = hist['gid'].astype(str).str.strip()
        
        gids_activos = df_ne['GID'].dropna().astype(str).str.strip().unique()
        
        if len(gids_activos) == 0:
            detalle = pd.DataFrame(columns=["GID", "Dias", "Circuito", "Posición", "Dirección"])
            return resumen, detalle
            
        hist = hist[hist['gid'].isin(gids_activos)].copy()
        hist = hist.sort_values(by=['gid', 'Fecha_hora_pasaje'])
        
        NO_ESTA_KEYS_NORM = ["NO ESTA (20)", "CONTENEDOR NO ESTA", "NO ESTA"]
        hist['en_no_esta'] = hist['motivo_norm'].isin(NO_ESTA_KEYS_NORM)
        
        # Identificar episodios (cambios de bloque)
        hist['prev'] = hist.groupby('gid')['en_no_esta'].shift(1).fillna(False)
        hist['entran'] = (~hist['prev']) & hist['en_no_esta']
        hist['epi_id'] = hist.groupby('gid')['entran'].cumsum()
        
        # Buscar el último evento
        ult = hist.groupby('gid').tail(1).copy()
        ult_en = ult[ult['en_no_esta']]
        
        # Sacar el inicio del episodio activo actual
        inicio_epi = hist[hist['en_no_esta']].merge(
            ult_en[['gid', 'epi_id']], on=['gid', 'epi_id'], how='inner'
        )
        inicio_epi_agg = inicio_epi.groupby(['gid', 'epi_id'])['Fecha_hora_pasaje'].min().reset_index()
        inicio_epi_agg.columns = ['gid', 'epi_id', 'inicio_no_esta']
        
        now = datetime.now()
        inicio_epi_agg['duracion_dias'] = (now - inicio_epi_agg['inicio_no_esta']).dt.total_seconds() / 86400
        inicio_epi_agg['duracion_dias'] = inicio_epi_agg['duracion_dias'].round(0).fillna(0).astype(int)
        
        # Unir con los atributos de foto del dia
        snap_attrs = df_ne[['GID', 'Circuito', 'Posición', 'Ubicación']].drop_duplicates(subset=['GID'])
        snap_attrs['GID'] = snap_attrs['GID'].astype(str).str.strip()
        
        detalle = inicio_epi_agg[['gid', 'duracion_dias']].merge(
            snap_attrs, left_on='gid', right_on='GID', how='left'
        )
        detalle = detalle[['GID', 'duracion_dias', 'Circuito', 'Posición', 'Ubicación']]
        detalle.columns = ["GID", "Dias", "Circuito", "Posición", "Dirección"]
        detalle = detalle.sort_values(by='Dias', ascending=False)
        
    except Exception as e:
        print(f"Error procesando historico no esta: {e}")
        detalle = pd.DataFrame(columns=["GID", "Dias", "Circuito", "Posición", "Dirección"])
        
    return resumen, detalle


def df_to_reportlab_table(df, title):
    if df.empty: return []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading2'], alignment=1, spaceAfter=12)
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], alignment=1, fontSize=8, leading=10)
    
    # Cabecera en negrita
    header_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], alignment=1, fontSize=9, leading=11, fontName='Helvetica-Bold')

    elements = [Paragraph(title, title_style)]
    
    # Headers
    headers = [Paragraph(str(col), header_style) for col in df.columns]
    data = [headers]
    
    # Llenamos celdas convirtiéndolas en Parrafos para que haga word-wrap automático
    for _, row in df.iterrows():
        row_data = []
        for val in row:
            text = str(val) if pd.notna(val) else ""
            row_data.append(Paragraph(text, cell_style))
        data.append(row_data)
        
    # Ancho de columnas adaptativo (A4 Portrait es 595.27 x 841.89 puntos) -> ~520 puntos usables
    base_w = 520 / len(df.columns)
    col_widths = [base_w] * len(df.columns)
    for i, col in enumerate(df.columns):
        if col in ['Tramos Posición', 'Dirección']:
            col_widths[i] = base_w * 1.8
        elif col in ['Cant.', 'Cantidad', 'Dias', '>=6', '5', '4', '3']:
            col_widths[i] = base_w * 0.6
            
    # Ajustar para que sumen exactamente 520
    factor = 500 / sum(col_widths)
    col_widths = [w * factor for w in col_widths]

    # Para que un listado muy largo se pase de página solito
    t = Table(data, colWidths=col_widths, repeatRows=1)
    
    t_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.lightgrey),
    ]
    
    # Pintar la fila total
    if len(df) > 0 and str(df.iloc[-1, 0]).upper() == 'TOTAL':
        t_style.append(('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#DCE6F1')))
        t_style.append(('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'))

    t.setStyle(TableStyle(t_style))
    elements.append(t)
    elements.append(Spacer(1, 20))
    elements.append(PageBreak())
    return elements

def export_to_pdf(tabla_final, detalle_at, resumen_grua, listado_grua, resumen_ne, resumen_fuego, detalle_ne, output_filename="Reporte_Operativa.pdf"):
    print(f"Exportando a {output_filename}...")
    base_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(base_dir, output_filename)
    
    doc = SimpleDocTemplate(out_path, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    story.extend(df_to_reportlab_table(tabla_final, "Resumen de Atrasos"))
    story.extend(df_to_reportlab_table(detalle_at, "Detalle de Atrasos por Circuito"))
    story.extend(df_to_reportlab_table(resumen_grua, "Resumen de Grúa"))
    story.extend(df_to_reportlab_table(listado_grua, "Listado Grúa"))
    story.extend(df_to_reportlab_table(resumen_ne, "Resumen de Contenedor no está"))
    story.extend(df_to_reportlab_table(resumen_fuego, "Resumen de Fuego"))
    if not detalle_ne.empty:
        story.extend(df_to_reportlab_table(detalle_ne, "Detalle No Esta"))
        
    doc.build(story)

def export_to_excel(dic_tablas, output_filename="Reporte_Operativa.xlsx"):
    print(f"Exportando a Excel {output_filename}...")
    base_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(base_dir, output_filename)
    
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        for sheet_name, df in dic_tablas.items():
            if df.empty:
                continue
                
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            
            # Definir estilos
            header_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
            header_font = Font(bold=True, name='Arial', size=11)
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Aplicar encabezados
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
                
            # Aplicar bordes y alineacion
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
                    col_name = df.columns[cell.column - 1]
                    if col_name in ['Tramos Posición', 'Dirección', 'Ubicación']:
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align
                        
            # Si tiene fila FINAL "TOTAL" la ponemos en negrita con fondo
            if not df.empty and str(df.iloc[-1, 0]).upper() == 'TOTAL':
                for cell in worksheet[worksheet.max_row]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
                    
            # Auto-ajustar celdas (Ancho maximo 60 caracteres)
            for i, col in enumerate(df.columns):
                max_length = len(str(col))
                for val in df[col]:
                    try:
                        if len(str(val)) > max_length:
                            max_length = len(str(val))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[get_column_letter(i+1)].width = adjusted_width

if __name__ == "__main__":
    print('Buscando archivo de datos...')
    try:
        _, dt_archivo = buscar_archivo_contenedores()
    except (FileNotFoundError, ValueError) as e:
        print(f"ERROR: {e}")
        exit(1)

    print(f"Archivo encontrado. Fecha/hora del archivo: {dt_archivo.strftime('%d/%m/%Y %H:%M')}")

    print('Cargando datos...')
    df, dt_archivo = load_data()
    if not df.empty:
        print('Procesando Atrasos...')
        tabla_final, detalle_at = procesar_atrasos(df)
        print('Procesando Grua...')
        res_grua, listado_grua = procesar_grua(df)
        print('Procesando Fuego...')
        res_fuego = procesar_fuego(df)
        print('Procesando No Esta...')
        res_ne, det_ne = procesar_no_esta(df)

        # Determinar el slot de hora según la hora del archivo
        # < 11:00 → 0915 | >= 12:00 → 1220
        hora_archivo = dt_archivo.hour
        if hora_archivo < 11:
            slot_hora = "0915"
        else:
            slot_hora = "1220"

        # Fecha en formato DD-MM-YY
        fecha_str = dt_archivo.strftime("%d-%m-%y")

        nombre_base = f"Reporte_Atrasos_Grua_No_esta_Fuego-{fecha_str}_{slot_hora}"
        pdf_out = f"{nombre_base}.pdf"
        excel_out = f"{nombre_base}.xlsx"

        export_to_pdf(tabla_final, detalle_at, res_grua, listado_grua, res_ne, res_fuego, det_ne, pdf_out)
        
        # Diccionario relacionando nombre de hoja Excel con dataframe procesado (Mismo orden que R)
        dic_excel = {
            "Resumen_Atrasos": tabla_final,
            "Detalle_Atrasos": detalle_at,
            "Resumen_Grua": res_grua,
            "Listado_Grua": listado_grua,
            "No_Esta": res_ne,
            "Fuego": res_fuego,
            "Detalle_no_esta": det_ne
        }
        export_to_excel(dic_excel, excel_out)
        
        print("Fin exitoso.")
    else:
        print("No se encontraron datos.")
